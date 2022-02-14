import mysql.connector as mysql
import os
import re
import requests
import smtplib

from asnake.client import ASnakeClient
from datetime import date
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from lxml import etree
from mysql.connector import errorcode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path
from secrets import *

id_field_regex = re.compile(r"(^id_+\d)")
id_combined_regex = re.compile(r'[\W_]+', re.UNICODE)
web_url_regex = re.compile(r"""(?i)\b((?:https?:(?:/{1,3}|[a-z0-9%])|[a-z0-9.\-]+[.](?:com|net|org|edu|gov|mil|aero|asia|biz|cat|coop|info|int|jobs|mobi|museum|name|post|pro|tel|travel|xxx|ac|ad|ae|af|ag|ai|al|am|an|ao|aq|ar|as|at|au|aw|ax|az|ba|bb|bd|be|bf|bg|bh|bi|bj|bm|bn|bo|br|bs|bt|bv|bw|by|bz|ca|cc|cd|cf|cg|ch|ci|ck|cl|cm|cn|co|cr|cs|cu|cv|cx|cy|cz|dd|de|dj|dk|dm|do|dz|ec|ee|eg|eh|er|es|et|eu|fi|fj|fk|fm|fo|fr|ga|gb|gd|ge|gf|gg|gh|gi|gl|gm|gn|gp|gq|gr|gs|gt|gu|gw|gy|hk|hm|hn|hr|ht|hu|id|ie|il|im|in|io|iq|ir|is|it|je|jm|jo|jp|ke|kg|kh|ki|km|kn|kp|kr|kw|ky|kz|la|lb|lc|li|lk|lr|ls|lt|lu|lv|ly|ma|mc|md|me|mg|mh|mk|ml|mm|mn|mo|mp|mq|mr|ms|mt|mu|mv|mw|mx|my|mz|na|nc|ne|nf|ng|ni|nl|no|np|nr|nu|nz|om|pa|pe|pf|pg|ph|pk|pl|pm|pn|pr|ps|pt|pw|py|qa|re|ro|rs|ru|rw|sa|sb|sc|sd|se|sg|sh|si|sj|Ja|sk|sl|sm|sn|so|sr|ss|st|su|sv|sx|sy|sz|tc|td|tf|tg|th|tj|tk|tl|tm|tn|to|tp|tr|tt|tv|tw|tz|ua|ug|uk|us|uy|uz|va|vc|ve|vg|vi|vn|vu|wf|ws|ye|yt|yu|za|zm|zw)/)(?:[^\s()<>{}\[\]]+|\([^\s()]*?\([^\s()]+\)[^\s()]*?\)|\([^\s]+?\))+(?:\([^\s()]*?\([^\s()]+\)[^\s()]*?\)|\([^\s]+?\)|[^\s`!()\[\]{};:'".,<>?«»“”‘’])|(?:(?<!@)[a-z0-9]+(?:[.\-][a-z0-9]+)*[.](?:com|net|org|edu|gov|mil|aero|asia|biz|cat|coop|info|int|jobs|mobi|museum|name|post|pro|tel|travel|xxx|ac|ad|ae|af|ag|ai|al|am|an|ao|aq|ar|as|at|au|aw|ax|az|ba|bb|bd|be|bf|bg|bh|bi|bj|bm|bn|bo|br|bs|bt|bv|bw|by|bz|ca|cc|cd|cf|cg|ch|ci|ck|cl|cm|cn|co|cr|cs|cu|cv|cx|cy|cz|dd|de|dj|dk|dm|do|dz|ec|ee|eg|eh|er|es|et|eu|fi|fj|fk|fm|fo|fr|ga|gb|gd|ge|gf|gg|gh|gi|gl|gm|gn|gp|gq|gr|gs|gt|gu|gw|gy|hk|hm|hn|hr|ht|hu|id|ie|il|im|in|io|iq|ir|is|it|je|jm|jo|jp|ke|kg|kh|ki|km|kn|kp|kr|kw|ky|kz|la|lb|lc|li|lk|lr|ls|lt|lu|lv|ly|ma|mc|md|me|mg|mh|mk|ml|mm|mn|mo|mp|mq|mr|ms|mt|mu|mv|mw|mx|my|mz|na|nc|ne|nf|ng|ni|nl|no|np|nr|nu|nz|om|pa|pe|pf|pg|ph|pk|pl|pm|pn|pr|ps|pt|pw|py|qa|re|ro|rs|ru|rw|sa|sb|sc|sd|se|sg|sh|si|sj|Ja|sk|sl|sm|sn|so|sr|ss|st|su|sv|sx|sy|sz|tc|td|tf|tg|th|tj|tk|tl|tm|tn|to|tp|tr|tt|tv|tw|tz|ua|ug|uk|us|uy|uz|va|vc|ve|vg|vi|vn|vu|wf|ws|ye|yt|yu|za|zm|zw)\b/?(?!@)))""")


def connect_aspace_api():
    """
    Connects to the ArchivesSpace staging API

    Returns:
         client (ASnake.client object): the ArchivesSpace ASnake client for accessing and connecting to the API
    """

    client: ASnakeClient = ASnakeClient(baseurl=as_api, username=as_auditor_un, password=as_auditor_pw)
    client.authorize()
    return client


def connect_db():
    """
    Connects to the ArchivesSpace staging database with credentials provided in local secrets.py file

    Returns:
         staging_connect: The connection to the staging database
         staging_cursor: The cursor of results for the staging database
    """

    try:
        staging_connect = mysql.connect(user=as_dbstag_un,
                                        password=as_dbstag_pw,
                                        host=as_dbstag_host,
                                        database=as_dbstag_database,
                                        port=as_dbstag_port)
    except mysql.Error as error:
        if error.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            print("Something is wrong with your username or password")
        elif error.errno == errorcode.ER_BAD_DB_ERROR:
            print("Database does not exist")
        else:
            print(error)
    else:
        staging_cursor = staging_connect.cursor()
        return staging_connect, staging_cursor


def email_users(send_from, send_to, subject, message, files=None, server="localhost", port=25, use_tls=False):
    """Compose and send email with provided info and attachments.

    Args:
        send_from (str): from name
        send_to (list[str]): to name(s)
        subject (str): message title
        message (str): message body
        files (list[str]): list of file paths to be attached to email
        server (str): mail server host name
        port (int): port number
        use_tls (bool): use TLS mode
    """

    if files is None:
        files = []
    send_to_str = ""
    for person in send_to:
        send_to_str += person + ", "
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = send_to_str[:-2]  # COMMASPACE.join(send_to)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    msg.attach(MIMEText(message))

    for path in files:
        part = MIMEBase('application', "octet-stream")
        with open(path, 'rb') as file:
            part.set_payload(file.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',
                        'attachment; filename={}'.format(Path(path).name))
        msg.attach(part)

    smtp = smtplib.SMTP(server, port)
    if use_tls:
        smtp.starttls()
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.quit()


def query_database(connection, cursor, statement):
    """
    Runs a query on the database

    Args:
        connection: The MySQL connection to the database
        cursor: Allows us to iterate a set of rows returned by a query.
                Source: https://www.mysqltutorial.org/mysql-cursor/
        statement (str): The MySQL statement to run against the database

    Returns:
        worksheet (openpysl.worksheet): An openpyxl worksheet class
    """

    cursor.execute(statement)
    results = cursor.fetchall()
    cursor.close()
    connection.close()
    return results


def generate_spreadsheet():
    """
    Creates a new spreadsheet for the data audit output, distinguished by date appended to end of filename

    Returns:
        wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit
        data_worksheet (str): The filepath of the data audit worksheet
    """

    wb = Workbook()
    data_spreadsheet = f'data_audit_{str(date.today())}.xlsx'
    wb.save(data_spreadsheet)
    return wb, data_spreadsheet


def write_headers(wb, sheetname, headers):
    """
    Takes a list of strings and writes them to the top row for a sheet in the data audit spreadsheet

    Args:
        wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit
        sheetname (str): The name of sheet to be added to the data audit spreadsheet
        headers (list): List of strings to be headers on the top row of the sheet

    Returns:
        worksheet (openpysl.worksheet): An openpyxl worksheet class
    """

    worksheet = wb.create_sheet(sheetname)
    worksheet.title = sheetname
    header_index = 0
    for row in worksheet.iter_rows(min_row=1, max_col=len(headers)):
        for cell in row:
            worksheet[cell.coordinate] = headers[header_index]
            worksheet[cell.coordinate].font = Font(bold=True, underline='single')
            header_index += 1
    return worksheet


def standardize_resids(results):
    """
    Takes the results of a MySQL statement and changes the resource identifier to be more readable

    Args:
        results (list): The results returned from a MySQL query

    Returns:
        updated_results (list): Updated results list containing more readable resource identifiers
    """

    updated_results = []
    for result in list(results):
        new_result = list(result)
        res_id_list = new_result[1].strip('[').strip(']').replace('"', '').split(',')
        combined_id = ''
        for id_comp in res_id_list:
            if id_comp != 'null':
                combined_id += id_comp + '-'
        new_result[1] = combined_id[:-1]
        updated_results.append(new_result)
    return updated_results


def update_booleans(results):
    """
    Takes the results of a MySQL statement and changes any 0 and 1 values to False and True, respectively

    Args:
        results (list): The results returned from a MySQL query

    Returns:
        checked_results (list): Updated results list containing True, False in place of 0s, 1s
    """

    updated_results = []
    for result in list(results):
        single_row = list(result)
        list_index = 0
        for value in single_row:
            if value == 0:
                single_row[list_index] = False
            elif value == 1:
                single_row[list_index] = True
            list_index += 1
        updated_results.append(single_row)
    return updated_results


def run_query(wb, sheetname, headers, statement, resid=False, booleans=False):
    """
    Takes a MySQL statement to run against the ArchivesSpace database

    Args:
        wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit
        sheetname (str): The name of sheet to be added to the data audit spreadsheet
        headers (list): List of strings to be headers on the top row of the sheet
        statement (str): The MySQL statement to run against the ArchivesSpace database
        resid (bool): If a resource identifier is included in the results of the MySQL statement, run the
                                identifier through standardize_resids()
        booleans (bool): If booleans are included in the results of the MySQL statement, run the
                                   them through update_booleans() to change them from numbers to TRUE or FALSE

    Returns:
        None
    """

    print(f'Checking {sheetname}... ', flush=True, end='')
    worksheet = write_headers(wb, sheetname, headers)
    connection, cursor = connect_db()
    results = query_database(connection, cursor, statement)
    standardized_results = results
    if resid is True:
        standardized_results = standardize_resids(results)
    if booleans is True:
        standardized_results = update_booleans(standardized_results)
    for result in standardized_results:
        worksheet.append(result)
    print("Done")


def check_controlled_vocabs(wb, vocab, terms, terms_num):
    """
    Takes a standard list of controlled vocabulary terms and checks an ArchivesSpace instance for any that do not match
    and highlights them in the data audit spreadsheet

    Args:
        wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit
        vocab (str): The controlled vocabulary (extent types, instance types, etc.) to be checked. Also serves as the
                     name of sheet to be added to the data audit spreadsheet
        terms (list): The standard list of terms to check against the terms found in ArchivesSpace
        terms_num (int): The enumeration identifier for a controlled vocabulary in the ArchivesSpace database

    Returns:
        None
    """

    print(f'\nChecking controlled vocabularies: {vocab}... ')
    write_row_index = 2
    headers = [str(vocab), "Read Only?", "Suppressed?"]
    vocab_sheet = write_headers(wb, str(vocab), headers)
    statement = (f'SELECT CONVERT(ev.value using utf8) AS {vocab}, ev.readonly AS Read_Only, '
                 f'ev.suppressed AS Suppressed '
                 f'FROM enumeration_value AS ev '
                 f'WHERE enumeration_id = {terms_num}')
    connection, cursor = connect_db()
    standardize_results = update_booleans(query_database(connection, cursor, statement))
    for result in standardize_results:
        vocab_sheet.append(result)
        if result[0] not in terms:
            print(f'Term not in standard list: {result[0]}')
            for cell in vocab_sheet[f'{write_row_index}:{write_row_index}']:
                cell.fill = PatternFill(start_color='FFFF0000',
                                        end_color='FFFF0000',
                                        fill_type='solid')
        write_row_index += 1
    print("Done")


def check_creators(wb, as_client):
    """
    Iterates through all published resources in an ArchivesSpace instance and checks to see if it has a creator agent

    Args:
        wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit
        as_client (ASnake.client object): the ArchivesSpace ASnake client for accessing and connecting to the API

    Returns:
        None
    """

    print("Checking resources without creators...")
    headers = ["Repository", "Resource ID", "Publish?", "Creator"]
    creator_sheet = write_headers(wb, "Resources without Creators", headers)
    repos = as_client.get("repositories").json()
    for repo in repos:
        print(repo["name"] + "\n")
        repo_id = repo["uri"].split("/")[2]
        resources = as_client.get("repositories/{}/resources".format(repo_id), params={"all_ids": True}).json()
        for resource_id in resources:
            has_creator = False
            resource = as_client.get("repositories/{}/resources/{}".format(repo_id, resource_id))
            combined_id = ""
            for field, value in resource.json().items():
                id_match = id_field_regex.match(field)
                if id_match:
                    combined_id += value + "-"
            combined_id = combined_id[:-1]
            if resource.json()["publish"] is True:
                if resource.status_code == 200:
                    if "linked_agents" in resource.json():
                        linked_agents = resource.json()["linked_agents"]
                        for linked_agent in linked_agents:
                            if linked_agent["role"] == "creator":
                                has_creator = True
                        if has_creator is False:
                            creator_sheet.append([repo["name"], combined_id, resource.json()["publish"], "None"])
                            print(f'Repo: {repo["name"]}, Resource: {combined_id}, '
                                  f'Publish?: {resource.json()["publish"]}')
                    else:
                        creator_sheet.append([repo["name"], combined_id, resource.json()["publish"], "None"])
                        print(f'Repo: {repo["name"]}, Resource: {combined_id}, '
                              f'Publish?: {resource.json()["publish"]}')
        print("_" * 100)


def check_child_levels(top_child_uri, root_uri, top_level, top_child_title, as_client):
    """
    Iterates through an archival object's published children and returns the top child (parent) URI, title, and the
    levels of the children of the parent archival object

    Args:
        top_child_uri (str): The URI string of the parent archival object
        root_uri (str): The root URI for the archival object being iterated through
        top_level (bool): If the archival object is on the top level of a resource (c01)
        top_child_title (str): The title of the parent archival object
        as_client (ASnake.client object): the ArchivesSpace ASnake client for accessing and connecting to the API

    Returns:
        top_child_uri (str): The URI string of the parent archival object
        top_child_title (str): The title of the parent archival object
        levels (list): A list of all levels of the children of the parent archival object
    """

    levels = []
    if top_level is True:
        tl_tree = as_client.get(f'{top_child_uri}/tree/root').json()
        for waypoint_num, waypoint_info in tl_tree["precomputed_waypoints"][''].items():
            for child in waypoint_info:
                child_info = as_client.get(f'{child["uri"]}').json()
                if child["level"] not in levels and child_info["publish"] is True:
                    levels.append(child["level"])
    else:
        children = as_client.get(root_uri + "/tree/node", params={"node_uri": top_child_uri,
                                                                  "published_only": True}).json()
        if children["child_count"] != 0:
            for waypoint_num, waypoint_info in children["precomputed_waypoints"][children["uri"]].items():
                for child in waypoint_info:
                    if child["level"] not in levels:
                        levels.append(child["level"])
    if "file" in levels and "item" in levels and len(levels) == 2:
        levels.clear()
    if len(levels) > 1:
        return top_child_uri, top_child_title, levels


def get_top_children(tree_info, child_levels, root_uri, aspace_coll_id, as_client, top_level=False):
    """
    Iterates through a resource's published archival objects in ArchivesSpace and returns all published archival
    objects that have children

    Args:
        tree_info (dict): The archival object tree for a resource
        child_levels (dict): The archival objects with children for a resource
        root_uri (str): The root URI for the child being iterated through
        aspace_coll_id (str): The identifier for the resource of the archival objects being iterated through
        as_client (ASnake.client object): the ArchivesSpace ASnake client for accessing and connecting to the API
        top_level (bool): If the archival object is on the top level of a resource (c01)

    Returns:
        child_levels (dict): The archival objects on the same level of each other within a resource
    """

    if tree_info["child_count"] > 0 and tree_info["uri"] not in child_levels:
        child_levels[f"{tree_info['uri']}"] = (tree_info["title"], tree_info["child_count"], tree_info["level"],
                                               aspace_coll_id, top_level)
    if "precomputed_waypoints" in tree_info and tree_info["child_count"] > 0:
        if top_level is True:
            waypoint_key = ""
        else:
            waypoint_key = tree_info["uri"]
        for waypoint_num, waypoint_info in tree_info["precomputed_waypoints"][waypoint_key].items():
            for child in waypoint_info:
                child_info = as_client.get(f'{child["uri"]}').json()
                if child["child_count"] > 0 and child_info["publish"] is True:
                    child_levels[f'{child["uri"]}'] = (child["title"], child["child_count"], child["level"],
                                                       aspace_coll_id, False)
                    children = as_client.get(root_uri + "/tree/node", params={"node_uri": child["uri"],
                                                                              "published_only": True}).json()
                    get_top_children(children, child_levels, root_uri, aspace_coll_id, as_client, top_level=False)
    return child_levels


def check_res_levels(wb, as_client):
    """
    Iterates through all published resources in ArchivesSpace instance and checks to see if archival objects on the top
    level are labeled with different levels; ex. file and series on c01, item, file, and subseries on c01

    Used for consistent display on SCLFind website for finding aids. Having an archival object labeled 'file' on the
    same level as an archival object labeled 'series' looks weird and is not consistent. We want archival objects on the
    same level to have consistent labels (with the exception of file and item as they are very similar).

    Args:
        wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit
        as_client (ASnake.client object): the ArchivesSpace ASnake client for accessing and connecting to the API

    Returns:
        None
    """

    print("Checking children levels...")
    headers = ["Repository", "Resource ID", "Parent Title", "Parent URI", "Level Disparity"]
    reslevel_sheet = write_headers(wb, "Collection Level Checks", headers)
    repos = as_client.get("repositories").json()
    for repo in repos:
        print(repo["name"] + "\n")
        repo_id = repo["uri"].split("/")[2]
        resources = as_client.get("repositories/{}/resources".format(repo_id), params={"all_ids": True}).json()
        for resource_id in resources:
            child_levels = {}
            resource = as_client.get("repositories/{}/resources/{}".format(repo_id, resource_id))
            combined_id = ""
            for field, value in resource.json().items():
                id_match = id_field_regex.match(field)
                if id_match:
                    combined_id += value + "-"
            combined_id = combined_id[:-1]
            if resource.json()["publish"] is True:
                if resource.status_code == 200:
                    root_uri = f'/repositories/{repo_id}/resources/{resource_id}'
                    tree_info = as_client.get(f'{root_uri}/tree/root').json()
                    child_levels = get_top_children(tree_info, child_levels, root_uri, combined_id, as_client,
                                                    top_level=True)
                    for top_child_uri, top_child_info in child_levels.items():
                        top_child_uri, top_child_title, level_disparity = check_child_levels(top_child_uri, root_uri,
                                                                                             top_child_info[4],
                                                                                             top_child_info[0],
                                                                                             as_client) or \
                                                                          (None, None, None)
                        if level_disparity is not None:
                            reslevel_sheet.append([repo["name"], combined_id, top_child_title, top_child_uri,
                                                   str(level_disparity)])
                            print(f'Repo: {repo["name"]}, Resource: {combined_id}, Parent Title: {top_child_title}, '
                                  f'Parent URI: {top_child_uri}, Level Disparity: {level_disparity}')


def duplicate_subjects(wb):
    """
    Checks for duplicate subjects in ArchivesSpace

    Args:
        wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit

    Returns:
        None
    """

    print("Checking for duplicate subjects...", flush=True, end='')
    headers = ["Original Subject", "Original Subject ID", "Duplicate Subject", "Duplicate Subject ID"]
    statement = f'SELECT title, id FROM subject'
    check_duplicates(wb, headers, statement, "Duplicate Subjects", "/subjects/")
    print("Done")


def duplicate_agent_persons(wb):
    """
    Checks for duplicate agent persons in ArchivesSpace

    Args:
        wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit

    Returns:
        None
    """

    print("Checking for duplicate agents...", flush=True, end='')
    headers = ["Original Agent", "Original Agent ID", "Duplicate Agent", "Duplicate Agent ID"]
    statement = f'SELECT sort_name, agent_person_id FROM name_person'
    check_duplicates(wb, headers, statement, "Duplicate Agents", "/agents/people/")
    print("Done")


def check_duplicates(wb, headers, statement, sheetname, uri_string):
    """
    Takes an SQL query and checks for duplicate results from the query

    Args:
        wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit
        headers (list): List of strings to be headers on the top row of the sheet
        statement (str): MySQL statement to be executed
        sheetname (str): The name of the sheet to be added to the data audit spreadsheet
        uri_string (str): Part of the URI of the duplicates found in ArchivesSpace, ex. '/agents/people' or '/subjects/'

    Returns:
        None
    """

    write_row_index = 2
    vocab_sheet = write_headers(wb, sheetname, headers)
    connection, cursor = connect_db()
    originals = query_database(connection, cursor, statement)
    comparisons = originals
    total_duplicates = []
    for original in originals:
        count = 0
        matches = {}
        for comparing_object in comparisons:
            if comparing_object[0] == original[0]:
                count += 1
                if original[0] in matches:
                    matches[original[0]].append(comparing_object[0])
                    matches[original[0]].append(f'{uri_string}{comparing_object[1]}')
                else:
                    matches[original[0]] = [comparing_object[0], f'{uri_string}{comparing_object[1]}']
        if count > 1:
            if matches not in total_duplicates:
                total_duplicates.append(matches)
    for duplicate in total_duplicates:
        for duplicate_name, duplicate_values in duplicate.items():
            vocab_sheet.append(duplicate[duplicate_name])
            write_row_index += 1


def create_export_folder():
    """
    Creates a directory at the same level of the script (source_eads) for storing all exported EAD.xml files
    """

    try:
        current_directory = os.getcwd()
        for root, directories, files in os.walk(current_directory):
            if "source_eads" in directories:
                return str(Path(os.getcwd(), "source_eads"))
            else:
                raise Exception
    except Exception as source_ead_error:
        print(str(source_ead_error) + "\nNo source_eads folder found, creating new one...", end='', flush=True)
        current_directory = os.getcwd()
        folder = "source_eads"
        source_path = os.path.join(current_directory, folder)
        os.mkdir(source_path)
        print("{} folder created\n".format(folder))
        return str(Path(source_path))


def delete_export_folder():
    """
    Deletes the source_eads directory and all files within if it exists
    """

    source_eads_path = str(Path.joinpath(Path.cwd(), "source_eads"))
    if os.path.exists(source_eads_path):
        for filename in os.listdir(source_eads_path):
            filepath = str(os.path.join(source_eads_path, filename))
            os.remove(filepath)
        os.rmdir(str(Path.joinpath(Path.cwd(), "source_eads")))


def export_eads(wb, source_path, as_client):
    """
    Calls the ArchivesSpace API to export all published resources as EAD.xml files with the following parameters:
        "include_unpublished": False,
        "include_daos": True,
        "numbered_cs": True,
        "print_pdf": False,
        "ead3": False

    Creates a directory at the same level of the script (source_eads) for storing all the EAD.xml files

    Args:
        wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit
        source_path (str): The filepath of the file being checked for URLs
        as_client (ASnake.client object): the ArchivesSpace ASnake client for accessing and connecting to the API

    Returns:
        None
    """

    print("Checking for EAD export errors...")
    headers = ["Repository", "Resource ID", "Export Error"]
    checkexports_sheet = write_headers(wb, "Export Errors", headers)
    repos = as_client.get("repositories").json()
    for repo in repos:
        print(repo["name"] + "\n")
        repo_id = repo["uri"].split("/")[2]
        resources = as_client.get("repositories/{}/resources".format(repo_id), params={"all_ids": True}).json()
        for resource_id in resources:
            resource = as_client.get("repositories/{}/resources/{}".format(repo_id, resource_id))
            combined_id = ""
            for field, value in resource.json().items():
                id_match = id_field_regex.match(field)
                if id_match:
                    combined_id += value + "-"
            combined_id = combined_id[:-1]
            combined_aspace_id_clean = id_combined_regex.sub('', combined_id)
            if resource.status_code == 200:
                if resource.json()["publish"] is True:
                    try:
                        export_ead = as_client.get("repositories/{}/resource_descriptions/{}.xml".format(repo_id,
                                                                                                         resource_id),
                                                   params={"include_unpublished": False, "include_daos": True,
                                                           "numbered_cs": True, "print_pdf": False, "ead3": False})
                    except Exception as e:
                        checkexports_sheet.append([repo["name", combined_aspace_id_clean, str(e)]])
                    else:
                        filepath = str(Path(source_path, combined_aspace_id_clean)) + ".xml"
                        with open(filepath, "wb") as local_file:
                            local_file.write(export_ead.content)
                            local_file.close()
                            print("Exported: {}".format(combined_id))
            else:
                checkexports_sheet.append([repo["name"], combined_aspace_id_clean, resource.json()])


def check_urls(wb, source_path):
    """
    Iterates through an .xml file checking for URLs using regex

    Args:
        wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit
        source_path (str): The filepath of the file being checked for URLs

    Returns:
        None
    """

    print("Checking for URL errors...")
    headers = ["Repository", "Resource ID", "Parent Title", "URL", "URL Error Code"]
    checkurls_sheet = write_headers(wb, "URL Errors", headers)
    repo = None
    resid = None
    for file in os.listdir(source_path):
        print(file)
        tree = etree.parse(source_path + "/" + file)
        root = tree.getroot()
        for element in root.getiterator():
            element.tag = etree.QName(element).localname
        for element in root.findall(".//*"):
            if element.tag == "corpname":
                if element.getparent().tag == "repository":
                    repo = element.text
            if element.tag == "unitid":
                resid = element.text
            if element.tag == "extref" or element.tag == "dao":
                attributes = dict(element.attrib)
                for key, value in attributes.items():
                    if key == "{http://www.w3.org/1999/xlink}href":
                        # print(f'Testing URL from href: {value}')
                        res = bool(re.search(r"\s", value))  # check if there are spaces in the URL
                        if res:
                            checkurls_sheet.append([repo, resid, element.getparent().getparent().tag, value,
                                                    "URL contains spaces, PDF exports will fail"])
                        response = check_url(value)
                        if response:
                            checkurls_sheet.append([repo, resid, element.getparent().getparent().tag, value, response])
            else:
                element_words = str(element.text).split(" ")
                filtered_words = list(filter(None, element_words))
                for word in filtered_words:
                    clean_word = word.strip(",.;:`~()<>")
                    match = web_url_regex.match(clean_word)
                    if match:
                        # print(f'Testing URL from content: {match}')
                        response = check_url(clean_word)
                        if response:
                            checkurls_sheet.append([repo, resid, element.getparent().getparent().tag, clean_word,
                                                    response])


def check_url(url):
    """
    Takes a string input as a URL and checks the response code

    Args:
        url (str): The URL string to test the response

    Returns:
        response_code (int): The response code from requesting the URL
    """

    response_code = None
    try:
        response = requests.get(url, allow_redirects=True, timeout=30)
        if response.status_code != 200:
            response_code = str(response)
            print(response)
    except Exception as e:
        response_code = str(e)
        print(e)
    finally:
        return response_code


def run_audit(workbook, spreadsheet):
    """
    Calls a series of functions to run data audits on UGA's ArchivesSpace staging data with the API and MySQL database.
    It generates an excel spreadsheet found in the reports directory

    Args:
        workbook (openpyxl.Workbook): openpyxl Workbook class to use to edit the spreadsheet
        spreadsheet (str): filename of the spreadsheet

    Returns:
        None
    """

    aspace_client = connect_aspace_api()
    controlled_vocabs = {"Subject_Term_Type": [["cultural_context", "function", "genre_form", "geographic",
                                                "occupation", "style_period", "technique", "temporal", "topical",
                                                "uniform_title"], 54],
                         "Subject_Sources": [["aat", "lcsh", "local", "tgn", "lcnaf"], 23],
                         "Finding_Aid_Status_Terms": [["completed", "unprocessed", "in_process", "problem"], 21],
                         "Name_Sources": [["local", "naf", "ingest"], 4],
                         "Instance_Types": [["audio", "books", "digital_object", "graphic_materials", "maps",
                                             "microform", "mixed_materials", "moving_images", "electronic_records",
                                             "artifacts"], 22],
                         "Extent_Types": [["linear_feet", "box(es)", "item(s)", "gigabyte(s)", "pages", "volume(s)",
                                           "moving_image(s)", "interview(s)", "minutes", "folder(s)",
                                           "sound_recording(s)", "photograph(s)", "oversize_folders"], 14],
                         "Digital_Object_Types": [["cartographic", "mixed_materials", "moving_image",
                                                   "software_multimedia", "sound_recording", "still_image", "text"],
                                                  12],
                         "Container_Types": [["box", "folder", "oversized_box", "oversized_folder", "reel", "roll",
                                              "portfolio", "item", "volume", "physdesc", "electronic_records", "carton",
                                              "drawer", "cassette", "rr", "cs"], 16],
                         "Accession_Resource_Types": [["collection", "papers", "records"], 7]}
    cuid_statement = ('SELECT repo.name AS Repository, resource.identifier AS Resource_ID, ao.ref_id AS Ref_ID, '
                      'ao.title AS Archival_Object_Title, ao.component_id AS Component_Unique_Identifier '
                      'FROM archival_object AS ao '
                      'JOIN repository AS repo ON repo.id = ao.repo_id '
                      'JOIN resource ON resource.id = ao.root_record_id '
                      'WHERE component_id is not Null '
                      'AND resource.publish is True '
                      'AND ao.publish is True')
    tcnb_statement = ('SELECT repo.name AS Repository, resource.identifier AS Resource_ID, '
                      'ao.ref_id AS Linked_Archival_Object_REFID, ao.title AS Linked_Archival_Object_Title, '
                      'top_container.indicator, CONVERT(ev.value using utf8) AS Container_Type '
                      'FROM top_container '
                      'JOIN top_container_link_rlshp AS top_rlsh ON top_rlsh.top_container_id = top_container.id '
                      'JOIN sub_container ON top_rlsh.sub_container_id = sub_container.id '
                      'JOIN instance ON instance.id = sub_container.instance_id '
                      'JOIN archival_object AS ao ON ao.id = instance.archival_object_id '
                      'JOIN repository AS repo ON repo.id = ao.repo_id '
                      'JOIN resource ON resource.id = ao.root_record_id '
                      'JOIN enumeration_value AS ev ON ev.id = top_container.type_id '
                      'WHERE top_container.barcode is NULL AND ao.publish is True AND resource.publish is True')
    tcnind_statement = ('SELECT repo.name AS Repository, resource.identifier AS Resource_ID, '
                        'ao.ref_id AS Linked_Archival_Object_REFID, ao.title AS Linked_Archival_Object_Title, '
                        'top_container.indicator, CONVERT(ev.value using utf8) AS Container_Type '
                        'FROM top_container '
                        'JOIN top_container_link_rlshp AS top_rlsh ON top_rlsh.top_container_id = top_container.id '
                        'JOIN sub_container ON top_rlsh.sub_container_id = sub_container.id '
                        'JOIN instance ON instance.id = sub_container.instance_id '
                        'JOIN archival_object AS ao ON ao.id = instance.archival_object_id '
                        'JOIN repository AS repo ON repo.id = ao.repo_id '
                        'JOIN resource ON resource.id = ao.root_record_id '
                        'JOIN enumeration_value AS ev ON ev.id = top_container.type_id '
                        'WHERE top_container.indicator is NULL AND ao.publish is True AND resource.publish is True')
    users_statement = ('SELECT name, username, is_system_user AS System_Administrator, is_hidden_user AS Hidden_User '
                       'FROM user')
    aomtc_statement = ('SELECT repository.name AS Repository, instance.archival_object_id, ao.ref_id, ao.title, '
                       'COUNT(*) AS Top_Container_Count '
                       'FROM instance '
                       'JOIN archival_object AS ao ON ao.id = instance.archival_object_id '
                       'JOIN repository ON repository.id = ao.repo_id '
                       'WHERE ao.publish is True AND instance.instance_type_id != 349 '
                       'group by instance.archival_object_id HAVING count(archival_object_id) > 1')
    aomdo_statement = ('SELECT repository.name AS Repository, instance.archival_object_id, ao.ref_id, ao.title, '
                       'COUNT(*) AS Digital_Object_Count '
                       'FROM instance '
                       'JOIN archival_object AS ao ON ao.id = instance.archival_object_id '
                       'JOIN repository ON repository.id = ao.repo_id '
                       'WHERE ao.publish is True AND instance.instance_type_id = 349 '
                       'group by instance.archival_object_id HAVING count(archival_object_id) > 1')
    aocollevel_statement = ("SELECT repository.name AS Repository, resource.identifier AS Resource_ID, ao.title, "
                            "ao.ref_id, ev.value "
                            "FROM archival_object AS ao "
                            "JOIN resource ON ao.root_record_id = resource.id "
                            "JOIN repository ON ao.repo_id = repository.id "
                            "JOIN enumeration_value AS ev ON ao.level_id = ev.id "
                            "WHERE ao.parent_id is Null "
                            "AND ao.publish = 1 "
                            "AND resource.publish is True "
                            "AND ao.level_id = 889")
    eadid_statement = ("SELECT repository.name, resource.title, resource.identifier, resource.ead_id "
                       "FROM resource "
                       "JOIN repository ON repository.id = resource.repo_id "
                       "WHERE resource.ead_id is not NULL AND resource.publish is TRUE")
    queries = {"Component Unique Identifiers": [["Repository", "Resource ID", "RefID", "Archival Object Title",
                                                 "Component Unique Identifier"], cuid_statement, {"resids": True},
                                                {"booleans": False}],
               "Top Containers - No Barcodes": [["Repository", "Resource ID", "RefID", "Archival Object Title",
                                                 "Top Container Indicator", "Container Type"], tcnb_statement,
                                                {"resids": True}, {"booleans": False}],
               "Top Containers - No Indicator": [["Repository", "Resource ID", "RefID", "Archival Object Title",
                                                  "Top Container Indicator", "Container Type"], tcnind_statement,
                                                 {"resids": True}, {"booleans": False}],
               "Users": [["Name", "Username", "System Administrator?", "Hidden User?"], users_statement,
                         {"resids": False}, {"booleans": True}],
               "Arch Objs-Multiple Top Conts": [["Repository", "Archival Object ID", "RefID",
                                                 "Archival Object title", "Top Container Count"],
                                                aomtc_statement, {"resids": False}, {"booleans": False}],
               "Arch Objs-Multiple Dig Objs": [["Repository", "Archival Object ID", "RefID",
                                                "Archival Object Title", "Digital Object Count"],
                                               aomdo_statement, {"resids": False}, {"booleans": False}],
               "Arch Objs-Collection Level": [["Repository", "Resource ID", "Archival Object Title", "RefID",
                                               "Level"], aocollevel_statement, {"resids": True}, {"booleans": False}],
               "EAD-IDs": [["Repository", "Resource Title", "Resource ID", "EAD ID"], eadid_statement,
                           {"resids": True}, {"booleans": False}]}

    # for term, info in controlled_vocabs.items():
    #     check_controlled_vocabs(workbook, term, info[0], info[1])
    #
    # for query, info in queries.items():
    #     headers, sql_statement, resids, bools = info[0], info[1], info[2]["resids"], info[3]["booleans"]
    #     run_query(workbook, query, headers, sql_statement, resid=resids, booleans=bools)
    #
    duplicate_subjects(workbook)
    # duplicate_agent_persons(workbook)
    # check_creators(workbook, aspace_client)
    # check_res_levels(workbook, aspace_client)
    source_path = create_export_folder()
    # export_eads(workbook, source_path, aspace_client)
    # check_urls(workbook, source_path)

    # try:
    #     workbook.remove(workbook["Sheet"])
    # except Exception as e:
    #     print(e)

    workbook.save(spreadsheet)


def email_error(script_error):
    """
    Emails admin in case an error is generated when running the script

    Args:
        script_error (str): The error message to be included in the email

    Returns:
        None
    """

    error_message = f'Audit failed with error: {script_error}'
    email_users(cs_email, [cs_email], 'data_audit-FAIL', error_message, server=email_server)


def run_script():
    """
    Runs run_audit() and email_users() functions to run data audit and email users the generated spreadsheet
    """

    audit_workbook, spreadsheet_filename = generate_spreadsheet()
    spreadsheet_filepath = str(Path.joinpath(Path.cwd(), spreadsheet_filename))
    try:
        run_audit(audit_workbook, spreadsheet_filename)
    except Exception as e:
        email_error(e)
    else:
        try:
            message_sample = f'ArchivesSpace data audit generated. See attachment.'
            email_users(cs_email, [cs_email], f'{spreadsheet_filename}', message_sample,  # ks_email, rl_email
                        files=[spreadsheet_filepath], server=email_server)
        except Exception as e:
            email_error(e)
    finally:
        try:
            if os.path.exists(spreadsheet_filepath):
                os.remove(spreadsheet_filepath)
            delete_export_folder()
        except Exception as e:
            email_error(e)


run_script()
