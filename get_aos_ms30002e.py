# This script gets all of the archival objects for ms3000_2e

from secrets import *
from asnake.aspace import ASpace
from asnake.client import ASnakeClient

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import json

aspace = ASpace(baseurl=as_api, username=as_un, password=as_pw)
client = ASnakeClient(baseurl=as_api, username=as_un, password=as_pw)
client.authorize()

try:
    open_cache = open("data/PB_CACHE.json", "r")
    read_cache = open_cache.read()
    PB_CACHE = json.loads(read_cache)
    open_cache.close()
except Exception as e:
    PB_CACHE = {}
    print(e)

MAX_STALENESS = 200000


def is_fresh(cache_entry):
    now = datetime.now().timestamp()
    staleness = now - cache_entry['cache_timestamp']
    return staleness < MAX_STALENESS


def get_unique_key(url):
    return url


def make_request_using_cache(aspace_url=""):
    unique_ident = get_unique_key(aspace_url)
    if unique_ident in PB_CACHE:
        if is_fresh(PB_CACHE[unique_ident]):
            print("Getting cached data...")
            return PB_CACHE[unique_ident]["content"]
        else:
            print("found url, not fresh")
            print("Making a request for new data...")  # <asnake.client.web_client.ASnakeClient object at 0x039AFC58>/repositories/4/resources/4306/tree
            res_tree = client.get(f'/repositories/4/resources/4306/tree').json()
            PB_CACHE[unique_ident] = {"content": res_tree}
            PB_CACHE[unique_ident]['cache_timestamp'] = datetime.now().timestamp()
            fw = open("data/PB_CACHE.json", "w")
            dumped_json_cache = json.dumps(PB_CACHE)
            fw.write(dumped_json_cache)
            fw.close()
            return PB_CACHE[unique_ident]["content"]
    else:
        print("didnt find url")
        print("Making a request for new data...")
        res_tree = client.get(f'/repositories/4/resources/4306/tree').json()
        PB_CACHE[unique_ident] = {"content": res_tree}
        PB_CACHE[unique_ident]['cache_timestamp'] = datetime.now().timestamp()
        fw = open("data/PB_CACHE.json", "w")
        dumped_json_cache = json.dumps(PB_CACHE)
        fw.write(dumped_json_cache)
        fw.close()
        return PB_CACHE[unique_ident]["content"]


def generate_spreadsheet(name):
    wb = Workbook()
    data_spreadsheet = f'reports/{name}'
    wb.save(data_spreadsheet)
    return wb, data_spreadsheet


def write_headers(wb, sheetname, headers):
    worksheet = wb.create_sheet(sheetname)
    worksheet.title = sheetname
    header_index = 0
    for row in worksheet.iter_rows(min_row=1, max_col=len(headers)):
        for cell in row:
            worksheet[cell.coordinate] = headers[header_index]
            worksheet[cell.coordinate].font = Font(bold=True, underline='single')
            header_index += 1
    return worksheet


data_spreadsheet = 'ms_3000_aos.xlsx'
workbook, spreadsheet = generate_spreadsheet(data_spreadsheet)

headers = ["Title", "URI"]
ms3000_sheet = write_headers(workbook, "ms3000_2e_AOs", headers)

aspace_request = f'/repositories/4/resources/4306/tree/root'
ms3000_json = make_request_using_cache(aspace_request)

for child in ms3000_json["children"]:
    ms3000_sheet.append([child["title"], child["record_uri"]])

try:
    workbook.remove(workbook["Sheet"])
except Exception as e:
    print(e)

workbook.save(spreadsheet)
