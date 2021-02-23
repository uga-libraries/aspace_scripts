# import re
from openpyxl import load_workbook
from secrets import *
from asnake.client import ASnakeClient

# date_regex = re.compile(r"((?:(?:between|approximately) )*\d{4}(?:(?:-| and )\d{4})).*")


def correct_dates_type(agent, agent_type, client):
    expression = ""
    count = 0
    agent_data = client.get(f"agents/{agent_type}/{agent}").json()
    if "dates_of_existence" not in agent_data:
        print(agent_data, f"agents/{agent_type}/{agent}")
    else:
        for date in agent_data["dates_of_existence"]:
            if "begin" in date and "end" in date:
                expression += date["begin"] + "-" + date["end"]
                date["date_type"] = "range"
            elif "begin" not in date and "end" in date:
                expression += "-" + date["end"]
                date["date_type"] = "single"
            elif "begin" in date and "end" not in date:
                expression += date["begin"] + "-"
                date["date_type"] = "single"
            elif "expression" in date:
                expression = date["expression"]
        if expression:
            for name in agent_data["names"]:
                if name["is_display_name"] is True:
                    if "dates" not in name or not name["dates"]:  # if dates already exists or if it doesn't have any values
                        name["dates"] = expression
                        try:
                            update_agent = client.post(f"agents/{agent_type}/{agent}", json=agent_data).json()
                            if "error" in update_agent:
                                raise Exception
                            else:
                                print(f"{agent_data['title']};{name['dates']};"
                                      f"{agent_data['dates_of_existence'][0]['date_type']};"
                                      f"{agent_data['uri']}")
                                count += 1
                        except Exception as e:
                            print(f"ERROR;{agent_data['title']};"
                                  f"URI: {agent_data['uri']}\nError: {e}")
                    else:
                        print(f"{agent_data['title']};{name['dates']};"
                              f"{agent_data['dates_of_existence'][0]['date_type']};"
                              f"{agent_data['uri']}")
    return count


def get_agents(client):
    count = 0
    agent_types = ["people", "families", "corporate_entities"]
    for agent_type in agent_types:
        print("Agent Title;Agent Dates;Agent Type;ASpace URI")
        print(agent_type + "\n")
        agents = client.get(f"agents/{agent_type}", params={"all_ids": True}).json()
        for agent in agents:
            count += correct_dates_type(agent, agent_type, client)
        print("\n" + "-"*100)
    print(f"Updated a total of {str(count)} agents")


def update_agents_spreadsheet(client, excel_filepath):
    wb = load_workbook(excel_filepath)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            agent = client.get(row[3]).json()
            print(f"Converting Name Forms - Dates: {agent['names'][0]['dates']} > {row[1]};Converting Date Type: {agent['dates_of_existence'][0]['date_type']} > {row[2]}; ",  end='', flush=True)
            agent["names"][0]["dates"] = str(row[1])
            agent["dates_of_existence"][0]["date_type"] = str(row[2])
            agent_update = client.post(row[3], json=agent)
            print(f"Done. Response: {agent_update.json()}")
        except Exception as e:
            print(f"Couldn't update {row[0]}, URI: {row[3]}; Error: {e}")


asp_client = ASnakeClient(baseurl=as_api_stag, username=as_un, password=as_pw)
asp_client.authorize()
get_agents(asp_client)
spreadsheet_filepath = input("Spreadsheet filepath: ")
update_agents_spreadsheet(asp_client, spreadsheet_filepath)
