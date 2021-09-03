# This script updates ArchivesSpace containers from a spreadsheet
from openpyxl import load_workbook
from secrets import *
from asnake.aspace import ASpace
from asnake.client import ASnakeClient


aspace = ASpace(baseurl=as_api, username=as_un, password=as_pw)
client = ASnakeClient(baseurl=as_api, username=as_un, password=as_pw)
client.authorize()

resource_id = input("Enter ASpace URI: ")
excel_filepath = input("Enter full filepath for spreadsheet: ")
wb = load_workbook(excel_filepath)
sheet = wb.active
for row in sheet.iter_rows(min_row=2, values_only=True):
    archival_object = client.get(row[0]).json()
    print("Converting: {} > {} ... ".format(archival_object["instances"][0]["sub_container"]["indicator_2"], row[5]),
          end='', flush=True)
    archival_object["instances"][0]["sub_container"]["indicator_2"] = str(row[5])
    update_ao = client.post(row[0], json=archival_object)
    print("Done. Response: {}".format(update_ao.json()))
