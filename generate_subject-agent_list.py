# This script writes all subjects and agents to a spreadsheet with their title and URI. This script was made as the
# first step in a project to cleanup subjects and agents in ArchivesSpace. The final step is outlined in
# update_subjects_agents.py
from openpyxl import Workbook
from secrets import *
from asnake.aspace import ASpace
from asnake.client import ASnakeClient


def remove_period_subjects(client):
    arc_subjects = client.get("/subjects", params={"all_ids": True}).json()
    for subject in arc_subjects:
        subject_data = client.get(f"/subjects/{subject}").json()
        last_term = subject_data["terms"][-1]
        if last_term["term"][-1] == ".":
            last_term["term"] = last_term["term"][0:-1]
            update_subject = client.post(f"/subjects/{subject}", json=subject_data)
            print(update_subject.json(), subject_data["title"])


def write_headers(sheet):
    headers = ["Title", "URI", "Action", "Merge Into (Optional)"]
    header_index = 0
    for row in sheet.iter_rows(min_row=1, max_col=4):
        for cell in row:
            sheet[cell.coordinate] = headers[header_index]
            header_index += 1


def write_data(sheet, data_list):
    for data in data_list:
        sheet.append(data)


def get_subjects(client):
    subjects_list = []
    arc_subjects = client.get("/subjects", params={"all_ids": True}).json()
    for subject in arc_subjects:
        subject_data = client.get(f"/subjects/{subject}").json()
        subjects_list.append((subject_data["title"], subject_data["uri"]))
    return subjects_list


def get_agents(client):
    agents_list = []
    agent_types = ["people", "families", "corporate_entities"]
    for agent_type in agent_types:
        agents = client.get(f"agents/{agent_type}", params={"all_ids": True}).json()
        for agent in agents:
            agent_data = client.get(f"/agents/{agent_type}/{agent}").json()
            agents_list.append((agent_data["title"], agent_data["uri"]))
    return agents_list


# remove_period_subjects(client)

def run(client, subjects, agents, workbook):
    excel_file = "reports/subjects_agents.xlsx"
    total_subjects = []
    total_agents = []
    print("Getting subjects...", flush=True, end="")
    try:
        total_subjects = get_subjects(client)
        print(f" {len(total_subjects)} retrieved.")
    except Exception as e:
        print(f"An error occurred when getting subjects: {e}")
    print("Getting agents...", flush=True, end="")
    try:
        total_agents = get_agents(client)
        print(f" {len(total_agents)} retrieved.")
    except Exception as e:
        print(f"An error occurred when getting agents: {e}")
    print(f"Writing headers to {excel_file}...", flush=True, end="")
    try:
        write_headers(subjects)
        write_headers(agents)
        print("Done.")
    except Exception as e:
        print(f"An error occured when writing headers: {e}")
    print(f"Writing subjects to {excel_file}...", flush=True, end="")
    try:
        write_data(subjects, total_subjects)
        print("Done.")
    except Exception as e:
        print(f"An error occurred when writing subjects to {excel_file}: {e}")
    print(f"Writing agents to {excel_file}...", flush=True, end="")
    try:
        write_data(agents, total_agents)
        print("Done.")
    except Exception as e:
        print(f"An error occurred when writing agents to {excel_file}: {e}")
    print(f"Saving {excel_file}...", flush=True, end="")
    try:
        workbook.save(excel_file)
        print("Done.")
    except Exception as e:
        print(f"An error occured when saving {excel_file}: {e}")


if __name__ == '__main__':
    aspace = ASpace(baseurl=as_api_stag, username=as_un, password=as_pw)
    client = ASnakeClient(baseurl=as_api_stag, username=as_un, password=as_pw)
    client.authorize()
    wb = Workbook()
    wb.active()
    subjects = wb.create_sheet("Subjects")
    subjects.title = "Subjects"
    agents = wb.create_sheet("Agents")
    agents.title = "Agents"
    sheet = wb.get_sheet_by_name("Sheet")
    wb.remove(sheet)
    run(client, subjects, agents, wb)
