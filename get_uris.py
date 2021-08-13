import PySimpleGUI as psg
import sys
import json
import platform
import subprocess
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from asnake.client import ASnakeClient


def gui():
    defaults = psg.UserSettings()
    close_program, client, repositories = get_aspace_log(defaults)
    if close_program is True:
        sys.exit()
    main_layout = [[psg.Text("Choose your repository:", font=("Roboto", 12))],
                   [psg.DropDown([repo for repo in repositories.keys()], readonly=True,
                                 default_value=defaults["repo_default"], key="_REPO_SELECT_",),
                    psg.Button(" SAVE ", key="_SAVE_REPO_")],
                   [psg.FileBrowse(' Select Digital Objects File ', file_types=(("Excel Files", "*.xlsx"),),),
                    psg.InputText(default_text=defaults['_DO_FILE_'], key='_DO_FILE_')],
                   [psg.FileBrowse(' Select Template ', file_types=(("Excel Files", "*.xlsx"),),),
                    psg.InputText(default_text=defaults['_DOTEMP_FILE_'], key='_DOTEMP_FILE_')],
                   [psg.Button(' START ', key='_WRITE_DOS_', disabled=False)],
                   [psg.Output(size=(80, 18), key="_output_")],
                   [psg.Button(" Open DO Template File ", key="_OPEN_DOTEMP_")]]
    window = psg.Window('Write Digital Objects to Template', layout=main_layout)
    while True:
        event, values = window.read()
        if event == psg.WINDOW_CLOSED or event == 'Exit':
            break
        if event == '_WRITE_DOS_':
            if not values['_DO_FILE_']:
                psg.popup_error('ERROR\nPlease select a digital objects file', font=("Roboto", 14), keep_on_top=True)
            elif not values['_DOTEMP_FILE_']:
                psg.popup_error('ERROR\nPlease select a digital object template', font=("Roboto", 14), keep_on_top=True)
            else:
                defaults['_DO_FILE_'] = values['_DO_FILE_']
                defaults['_DOTEMP_FILE_'] = values['_DOTEMP_FILE_']
                write_digobjs(values['_DO_FILE_'], values['_DOTEMP_FILE_'], client,
                                       repositories[values["_REPO_SELECT_"]], window)
        if event == "_SAVE_REPO_":
            defaults["repo_default"] = values["_REPO_SELECT_"]
        if event == "_OPEN_DOTEMP_":
            if not defaults["_DOTEMP_FILE_"]:
                filepath_eads = str(Path.cwd())
                open_file(filepath_eads)
            else:
                filepath_eads = str(Path(defaults["_DOTEMP_FILE_"]))
                open_file(filepath_eads)


def write_digobjs(digobj_file, dotemp_file, client, repo, gui_window):
    gui_window[f'{"_WRITE_DOS_"}'].update(disabled=True)
    digobj_wb = load_workbook(digobj_file)
    digobj_sheet = digobj_wb.active
    dotemp_wb = load_workbook(dotemp_file)
    dotemp_sheet = dotemp_wb.active
    write_row_index = 6
    total_digobjs = 0
    digobj_columns = {0: "digital_object_id", 2: "digital_object_title", 3: "file_version_file_uri",
                      5: "date_1_expression", 8: "digital_object_publish"}
    sheet_columns = {}
    sheet_colnum = 0
    for col in digobj_sheet.iter_cols(max_col=digobj_sheet.max_column, values_only=True):
        sheet_columns[sheet_colnum] = col[0]
        sheet_colnum += 1
    for col_key, col_value in digobj_columns.items():
        if sheet_columns[col_key] != col_value:
            psg.popup_error("ERROR:\nDigital Object file columns do not match!\n\n"
                            "Check to make sure you entered the correct file")
            print("ERROR: Digital Object file columns do not match!\n\nCheck to make sure you entered the correct file")
            close_wbs(digobj_wb, dotemp_wb)
            gui_window[f'{"_WRITE_DOS_"}'].update(disabled=False)
            return
    errors = []
    for row in digobj_sheet.iter_rows(min_row=2, values_only=True):
        write_row_index += 1
        total_digobjs += 1
        digobj_id = row[0]
        digobj_title = row[2]
        digobj_url = row[3]
        digobj_date = row[5]
        digobj_publish = row[8]
        archobj_uri, resource_uri = get_results(client, repo, digobj_title, digobj_date)
        if archobj_uri is None and resource_uri is None:
            archobj_uri = "!!ERROR!!"
            resource_uri = "!!ERROR!!"
            errors.append(f'{digobj_title}, {digobj_date}')
            for cell in dotemp_sheet[f'{write_row_index}:{write_row_index}']:
                cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        write_obj_error = write_digobj(resource_uri, archobj_uri, digobj_id, digobj_title, digobj_publish,
                                       digobj_url, dotemp_sheet, write_row_index, dotemp_wb, dotemp_file,
                                       gui_window)
        if write_obj_error is not None:
            print(write_obj_error)
            close_wbs(digobj_wb, dotemp_wb)
        else:
            print(f'Written: {digobj_title}, {digobj_date}')
    close_wbs(digobj_wb, dotemp_wb)
    if errors:
        error_message = "ERROR: Could not find any records with the following titles:\n"
        for error in errors:
            error_message += "\n" + error + "\n"
        print(error_message)
    print(f'\n{"*" * 112}\n{" " * 40}Finished writing {total_digobjs} to {dotemp_sheet}\n{"*" * 112}')
    gui_window[f'{"_WRITE_DOS_"}'].update(disabled=False)
    return errors


def close_wbs(digobj_wb, dotemp_wb):
    digobj_wb.close()
    dotemp_wb.close()


def get_results(client, repo, digobj_title, digobj_date):
    archobj_uri = None
    resource_uri = None
    search_archobjs = client.get_paged(f"/repositories/{repo}/search",
                                       params={"q": f'title:"{digobj_title}"',  # , {digobj_date}
                                               "type": ['archival_object']})
    search_results = []
    for results in search_archobjs:
        search_results.append(results)
    if len(search_results) > 1:
        search_options = []
        for result in search_results:
            result_container_uri = result["top_container_uri_u_sstr"][0]
            top_container_json = client.get(result_container_uri).json()
            box_coll_info = top_container_json["long_display_string"]
            box_coll_list = box_coll_info.split(",")
            result_child = result["child_container_u_sstr"][0]
            result_option = f'{result["title"]}; {box_coll_list[0]}; {result_child}; {box_coll_list[1]}'
            search_options.append(result_option)
        multresults_layout = [[psg.Text(f'\n\nFound multiple options for\n{digobj_title}, {digobj_date}\n\n'  # TODO: add Box and Folder #, and identifier (ms1000) for info
                                        f'Choose one of the following:\n')],
                              [psg.Listbox(search_options, size=(120, 5),
                                           key="_ARCHOBJ_FILE_")],
                              [psg.Button(" SELECT ", key="_SELECT_ARCHOBJ_")]]
        multresults_window = psg.Window("Multiple Results for Archival Object", multresults_layout)
        selection = True
        while selection is True:
            multresults_event, multresults_values = multresults_window.Read()
            if multresults_event == "_SELECT_ARCHOBJ_":
                for result in search_results:
                    selection_title = multresults_values["_ARCHOBJ_FILE_"][0].split(";")[0]  # TODO: this part is screwing with me - not matching on second option at all, despite being perflectly matched
                    print(selection_title)
                    if result["title"].split(";")[0] == selection_title:
                        archobj_uri = result["uri"]
                        resource_uri = result["resource"]
                        selection = False
                        multresults_window.close()
                        break
                    else:
                        psg.popup_error("ERROR\nSelected result does not match!", font=("Roboto", 14),
                                        keep_on_top=True)
    elif len(search_results) == 0:
        print(f'\nERROR: No results found for:\n{digobj_title}, {digobj_date}\n')
        return archobj_uri, resource_uri
    else:
        for result in search_results:
            archobj_uri = result["uri"]
            resource_uri = result["resource"]
    if archobj_uri is None and resource_uri is None:
        print(f'{digobj_title}, {digobj_date}')
        print(search_results)
    return archobj_uri, resource_uri


def write_digobj(resource_uri, archobj_uri, digobj_id, digobj_title, digobj_publish, digobj_url, dotemp_sheet,
                 write_row_index, dotemp_wb, dotemp_file, gui_window):
    column_map = {4: resource_uri, 6: archobj_uri, 7: digobj_id, 8: digobj_title, 9: digobj_publish, 10: digobj_url}
    for column_num, column_value in column_map.items():
        dotemp_sheet.cell(row=write_row_index, column=column_num).value = column_value
    write_row_index += 1
    # print(f'Wrote {digobj_title} to {dotemp_file}\n')
    try:
        dotemp_wb.save(dotemp_file)
        return None
    except Exception as e:
        error = f'\n\nFailed opening {dotemp_file}. Please close the record before trying again.\nError: {e}'
        gui_window[f'{"_WRITE_DOS_"}'].update(disabled=False)
        return error


def get_aspace_log(defaults):
    """
    Gets a user's ArchiveSpace credentials.
    There are 3 components to it, the setup code, correct_creds while loop, and the window_asplog_active while loop. It
    uses ASnake.client to authenticate and stay connected to ArchivesSpace. Documentation for ASnake can be found here:
    https://archivesspace-labs.github.io/ArchivesSnake/html/index.html
    Args:
        defaults (UserSetting class): contains data from defaults.json file, all data the user has specified as default
    Returns:
        close_program (bool): if a user exits the popup, this will return true and end run_gui()
        connect_client (ASnake.client object): the ArchivesSpace ASnake client for accessing and connecting to the API
    """
    connect_client = None
    repositories = {}
    save_button_asp = " Save and Continue "
    window_asplog_active = True
    correct_creds = False
    close_program = False
    while correct_creds is False:
        asplog_col1 = [[psg.Text("ArchivesSpace username:", font=("Roboto", 11))],
                       [psg.Text("ArchivesSpace password:", font=("Roboto", 11))],
                       [psg.Text("ArchivesSpace API URL:", font=("Roboto", 11))]]
        asplog_col2 = [[psg.InputText(focus=True, key="_ASPACE_UNAME_")],
                       [psg.InputText(password_char='*', key="_ASPACE_PWORD_")],
                       [psg.InputText(defaults["as_api"], key="_ASPACE_API_")]]
        layout_asplog = [
            [psg.Column(asplog_col1, key="_ASPLOG_COL1_", visible=True),
             psg.Column(asplog_col2, key="_ASPLOG_COL2_", visible=True)],
            [psg.Button(save_button_asp, bind_return_key=True, key="_SAVE_CLOSE_LOGIN_")]
        ]
        window_login = psg.Window("ArchivesSpace Login Credentials", layout_asplog)
        while window_asplog_active is True:
            event_log, values_log = window_login.Read()
            if event_log == "_SAVE_CLOSE_LOGIN_":
                try:
                    connect_client = ASnakeClient(baseurl=values_log["_ASPACE_API_"],
                                                  username=values_log["_ASPACE_UNAME_"],
                                                  password=values_log["_ASPACE_PWORD_"])
                    connect_client.authorize()
                    defaults["as_api"] = values_log["_ASPACE_API_"]
                    repo_results = connect_client.get('/repositories')
                    repo_results_dec = json.loads(repo_results.content.decode())
                    for result in repo_results_dec:
                        uri_components = result["uri"].split("/")
                        repositories[result["name"]] = int(uri_components[-1])
                    window_asplog_active = False
                    correct_creds = True
                except Exception as e:
                    error_message = ""
                    if ":" in str(e):
                        error_divided = str(e).split(":")
                        for line in error_divided:
                            error_message += line + "\n"
                    else:
                        error_message = str(e)
                    psg.Popup("Your username and/or password were entered incorrectly. Please try again.\n\n" +
                              error_message)
            if event_log is None or event_log == 'Cancel':
                window_login.close()
                window_asplog_active = False
                correct_creds = True
                close_program = True
                break
        window_login.close()
    return close_program, connect_client, repositories


def open_file(filepath):
    """
    Takes a filepath and opens the folder according to Windows, Mac, or Linux.

    Args:
        filepath (str): the filepath of the folder/directory a user wants to open

    Returns:
        None
    """
    if platform.system() == "Windows":
        os.startfile(filepath)
    elif platform.system() == "Darwin":
        subprocess.Popen(["open", filepath])
    else:
        subprocess.Popen(["xdg-open", filepath])


if __name__ == "__main__":
    gui()
