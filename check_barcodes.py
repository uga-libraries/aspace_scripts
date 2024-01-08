# Checks ArchivesSpace against exported barcodes from our container management system to determine which barcodes do
# not exist in ArchivesSpace. Returns a csv of barcodes not found in ArchivesSpace
import csv
from secrets import *
from openpyxl import load_workbook
from asnake.client import ASnakeClient

client = ASnakeClient(baseurl=as_api, username=as_un, password=as_pw)
client.authorize()

barcode_wb = load_workbook("data/RLBC.xlsx")
barcode_sheet = barcode_wb["Item"]
count = 0
missing_barcodes = []
for row in barcode_sheet.iter_rows(values_only=True):
    count += 1
    search_barcodes = client.get_paged('/search', params={"q": 'barcode:' + row[0], "type": ['top_container']})
    search_results = [result for result in search_barcodes]
    if not search_results:
        print(f'Barcode {row[0]} not found')
        missing_barcodes.append(row[0])
    for result in search_results:
        if "barcode" in result:
            if row[0] != result["barcode"]:
                print(f'Search result does not match barcode {row[0]}: {result}')

with open('data/missing_barcodes-RLBC-tc.csv', mode='w', encoding='utf8', newline="") as missing_file:
    missing_file = csv.writer(missing_file, delimiter=',')
    for barcode in missing_barcodes:
        missing_file.writerow([barcode])

