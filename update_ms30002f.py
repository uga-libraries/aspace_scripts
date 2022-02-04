# This script transfers a series of archival objects from ms3000_2e to ms3000_2f at the top level

from secrets import *
from asnake.client import ASnakeClient
from openpyxl import load_workbook
from pathlib import Path

client = ASnakeClient(baseurl=as_api, username=as_un, password=as_pw)
client.authorize()

ao_workbook = load_workbook(Path.joinpath(Path.cwd(), "data", "MS3000_2fURI.xlsx"))
ao_sheet = ao_workbook["ms3000_2e_AOs"]
aobjects = []
for row in ao_sheet.iter_rows(min_col=1, min_row=2, values_only=True):
    arch_obj, target_res = row[0], row[1]
    transfer_ao = client.post("repositories/4/component_transfers",
                              params={"target_resource": target_res,
                                      "component": arch_obj}).json()
    print(transfer_ao)
