# This script grabs subjects from the ArchivesSpace database and their links to all resources and generates a
# spreadsheet with that info
import csv
import mysql.connector
import pathlib
from openpyxl import load_workbook
from secrets import sql_host, sql_user, sql_passwd

subjects = load_workbook(pathlib.Path.joinpath(pathlib.Path.cwd(), "reports", "subjects_agents-notes.xlsx"))
subjects_to_agents = subjects["Subjects to agents"]
subject_ids = []
for row in subjects_to_agents.iter_rows(min_col=2, values_only=True):
    for cell in row:
        subject_ids.append(cell.split("/")[-1])

uga_aspace_staging = mysql.connector.connect(host=sql_host, user=sql_user, password=sql_passwd,
                                             database="archivesspace")
astag_cursor = uga_aspace_staging.cursor(buffered=True)

for subject_id in subject_ids:
    collections = [subject_id]
    subres_query = f"SELECT subject.title, resource.identifier, repository.name " \
                   f"FROM subject_rlshp " \
                   f"JOIN subject ON subject.id = subject_rlshp.subject_id " \
                   f"JOIN resource ON resource.id = subject_rlshp.resource_id " \
                   f"JOIN repository ON repository.id = resource.repo_id " \
                   f"WHERE subject.id = {subject_id}"
    astag_cursor.execute(subres_query)
    results = astag_cursor.fetchall()
    for result in results:
        if result[0] not in collections:
            collections.append(result[0])
        collection = ""
        for collection_id_frag in list(result[1].strip("][").replace('"', '').split(",")):
            if collection_id_frag != "null":
                collection += collection_id_frag + "-"
        collection = collection[:-1]
        collections.append(collection)
    with open(pathlib.Path.joinpath(pathlib.Path.cwd(), "reports", "subject_to_agents.csv"), mode="a", newline="") as sub_to_ag:
        subag_writer = csv.writer(sub_to_ag, delimiter=",")
        subag_writer.writerow(collections)
