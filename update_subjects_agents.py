from openpyxl import load_workbook
from secrets import *
from asnake.client import ASnakeClient

client = ASnakeClient(baseurl=as_api_stag, username=as_un, password=as_pw)
client.authorize()

subject_wb = load_workbook("data/subjects_agents-notes.xlsx")
upsub_sheet = subject_wb["subjects_agents xlsx"]
for row in upsub_sheet.iter_rows(min_row=2, values_only=True):
    subject_title = row[0]
    subject_uri = row[4]
    subject_action = row[5]
    subject_merge = row[6]
    if subject_action == "DELETE":
        delete_response = client.delete(subject_uri)
        if "error" in delete_response.json():
            error_response = delete_response.json()
            print(f'ERROR when deleting {subject_title}: {error_response["error"]}')
        else:
            print(f'DELETED {subject_title}; Response: {delete_response.json()}')
    elif subject_action == "MERGE":
        updated_json = {'uri': 'merge_requests/subject',
                        'target': {'ref': str(subject_merge)},
                        'victims': [{'ref': str(subject_uri)}]}
        merge_response = client.post('merge_requests/subject', json=updated_json)
        if "error" in merge_response.json():
            error_response = merge_response.json()
            print(f'ERROR when merging {subject_title}: {error_response["error"]}')
        else:
            print(f'MERGED {subject_title} into {subject_merge}; Response: {merge_response.json()}')
