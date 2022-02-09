# Gets all archival objects in collections ms3000_1a, ms3000_1b, ms3000_2a, and ms3000_2b that don't have folders in
# their container instances

from openpyxl import Workbook
from openpyxl.styles import Font
from secrets import *
from asnake.client import ASnakeClient

client = ASnakeClient(baseurl=as_api, username=as_un, password=as_pw)
client.authorize()

resources = {'ms3000_1a': '/repositories/4/resources/4048', 'ms3000_1b': '/repositories/4/resources/4064',
             'ms3000_2a': '/repositories/4/resources/4049', 'ms3000_2b': '/repositories/4/resources/4059'}


def generate_spreadsheet():
    """
    Creates a new spreadsheet for the data audit output, distinguished by date appended to end of filename

    Returns:
        wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit
        data_worksheet (str): The filepath of the data audit worksheet
    """
    wb = Workbook()
    data_spreadsheet = f'reports/ms3000_nofolders.xlsx'
    wb.save(data_spreadsheet)
    return wb, data_spreadsheet


def write_headers(wb, sheetname, headers):
    """
    Takes a list of strings and writes them to the top row for a sheet in the data audit spreadsheet

    Args:
        wb (openpyxl.Workbook): The openpyxl workbook of the spreadsheet being generated for the data audit
        sheetname (str): The name of sheet to be added to the data audit spreadsheet
        headers (list): List of strings to be headers on the top row of the sheet

    Returns:
        worksheet (openpysl.worksheet): An openpyxl worksheet class
    """
    worksheet = wb.create_sheet(sheetname)
    worksheet.title = sheetname
    header_index = 0
    for row in worksheet.iter_rows(min_row=1, max_col=len(headers)):
        for cell in row:
            worksheet[cell.coordinate] = headers[header_index]
            worksheet[cell.coordinate].font = Font(bold=True, underline='single')
            header_index += 1
    return worksheet


def get_aos(tree_info, aosnc, root_uri, aspace_coll_id, as_client, top_level=False):
    if top_level is True:
        for child in tree_info:
            if "containers" in child:
                for container in child["containers"]:
                    if "type_2" not in container:
                        aosnc[f'{child["uri"]}'] = (child["title"], child["level"],
                                                    f'{container["top_container_type"]} '
                                                    f'{container["top_container_indicator"]}',
                                                    container["top_container_barcode"],
                                                    aspace_coll_id)
            if child["child_count"] > 0:
                children = as_client.get(root_uri + "/tree/node", params={"node_uri": child["uri"],
                                                                          "published_only": True}).json()
                get_aos(children, aosnc, root_uri, aspace_coll_id, as_client, top_level=False)
    else:
        child_tree = client.get(f'{root_uri}/tree/node', params={"node_uri": tree_info["uri"]}).json()
        waypoints = child_tree["waypoints"]
        for waypoint_page in range(0, waypoints):
            waypoint_tree = client.get(f'{root_uri}/tree/waypoint', params={'offset': waypoint_page,
                                                                            "parent_node": child_tree["uri"]}).json()
            get_aos(waypoint_tree, aosnc, tree_info["uri"], aspace_coll_id, client,
                    top_level=True)
    return aosnc


def run():
    workbook, spreadsheet = generate_spreadsheet()
    for coll_id, uri in resources.items():
        aos_no_child = {}
        headers = ["Archival Object Title", "URI", "Level", "Top Container", "Barcode"]
        nochildinst_sheet = write_headers(workbook, f'{coll_id}', headers)
        res_tree = client.get(f'{uri}/tree/root').json()
        waypoints = res_tree["waypoints"]
        for waypoint_page in range(0, waypoints):
            waypoint_tree = client.get(f'{uri}/tree/waypoint', params={'offset': waypoint_page}).json()
            aos_no_child = get_aos(waypoint_tree, aos_no_child, uri, coll_id, client, top_level=True)
        if aos_no_child:
            for ao_uri, ao_info in aos_no_child.items():
                nochildinst_sheet.append([ao_info[0], ao_uri, ao_info[1], ao_info[2], ao_info[3]])
    try:
        workbook.remove(workbook["Sheet"])
    except Exception as e:
        print(e)

    workbook.save(spreadsheet)


run()
